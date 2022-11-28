import os
os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE"
import random
from Environments.Demand import Demand
from Environments.SupplyLoss import SupplyLoss
from Agents.Retailer import Retailer
from Agents.Parent import Parent
from Actions.Action_retailer import Action_retailer
from Actions.Action_parent import Action_parent
from States.States import States
from Simulator.Sim import Sim
import time
from tqdm import tqdm
import numpy as np
from matplotlib import pyplot as plt
from Actions.X_retailer import X_retailer
from Simulator.NNModel import MLPModel
from openpyxl import Workbook
from datetime import datetime


if __name__ == '__main__':
    
    # Default setting
    maxEpisode = 100    # 100
    maxT = 10000   # number of periods, 10000, for testing
    
    
    
    initialInventoryLevel_1 = 30
    initialInventoryLevel_2 = 30
    
    # demand mean
    demandMeanMin_1 = 10
    demandMeanMax_1 = 60
    demandMeanMin_2 = 10
    demandMeanMax_2 = 60
    
    
    # demand std
    demandStdMin_1 = 1
    demandStdMax_1 = 5 
    demandStdMin_2 = 1
    demandStdMax_2 = 5 


    # supply mean
    supplyLossMeanMin_1 = 0
    supplyLossMeanMax_1 = 10
    supplyLossMeanMin_2 = 0
    supplyLossMeanMax_2 = 10
    
    
    # supply std
    supplyLossStdMin_1 = 1
    supplyLossStdMax_1 = 3  #   
    supplyLossStdMin_2 = 1
    supplyLossStdMax_2 = 3  #  
    
    
    # base stock upper/lower bound
    baseStockUpperBound = 70
    baseStockLowerBound1 = 20   
    baseStockLowerBound2 = 20   
    
    
    # population size
    populationSize = 70
    
    
    # generation size
    generationSize = 200
    
    
    # Supplier's capacity
    supplierCapacity = 60
    
    
    # seed number for GA (training)
    seedCR1 = 209382
    seedCR2 = 109382
    seedMR1 = 809214
    seedMR2 = 809234
    seedOne1 = 901112
    seedOne2 = 901111
    seedMut1 = 141423
    seedMut2 = 141413
    seedInitialPop1 = 3828109
    seedInitialPop2 = 3128109
    seedMating1 = 777878
    seedMating2 = 111911


    
    ############################
    # Experiment Type
    ############################
    expType = 0 # 0: Unit Transshipment Revenue Tuning
                # 1: Unit Holding Cost Sensitivity Analysis
                # 3: Performance Evaluation
                # 4: For counting unvisited (state, action) pair
                # 5: For base stock level policy with Genetic Algorithm
    
    
    
    type = 5   
                # 1: Decentralized Q (trans X)
                # 3: Centralized Q (trans O)
                # 5: Hetero Maximax Q (trans O)
                # 6: Base stock level policy with GA (trans X)
    
    if type == 6:
        maxEpisode = 1
        
    
    # Risk Sensitivity
    riskSensitivity1 = 1
    riskSensitivity2 = 1
                

    # Seed number (training)
    seedDemandMean_1 = 324253 
    seedDemandMean_2 = 809832
    seedDemandStdDev_1 = 123456
    seedDemandStdDev_2 = 122556
    seedDemandRealization_1 = 654321
    seedDemandRealization_2 = 614311
    seedSupplyLossMean_1 = 342121
    seedSupplyLossMean_2 = 803312
    seedSupplyLossStdDev_1 = 324251
    seedSupplyLossStdDev_2 = 334454
    seedSupplyLossRealization_1 = 132435
    seedSupplyLossRealization_2 = 112135
    seedEGreedy = 987239
    seedE2Greedy = 423432
    seedPi = 523698
    seedA1 = 223698
    seedA2 = 421698
    seedX1 = 932822
    seedX2 = 784821
    seedO = 921698


    # Seed number (test)
    seedDemandMean_1_test = 325123
    seedDemandMean_2_test = 352832
    seedDemandStdDev_1_test = 112456
    seedDemandStdDev_2_test = 111156
    seedDemandRealization_1_test = 622221
    seedDemandRealization_2_test = 633311
    seedSupplyLossMean_1_test = 344421
    seedSupplyLossMean_2_test = 804442
    seedSupplyLossStdDev_1_test = 555251
    seedSupplyLossStdDev_2_test = 334224
    seedSupplyLossRealization_1_test = 134435
    seedSupplyLossRealization_2_test = 145535
    seedPi_test = 522228
    
    
    
    # Summary 
    avgAvgRewardList = []
    avgAvgUnmetDemandList = []
    avgAvgAmountTransTo1List = []
    avgAvgAmountTransTo2List = []
    avgAvgAmountOrder1List = []
    avgAvgAmountOrder2List = []
    avgAvgOnHandInventory1List = []
    avgAvgOnHandInventory2List = []
    avgTimeSecList = []
    
    
    
    
    ############################
    # Unit Transshipment Revenue
    ############################
    if expType == 0:
        # For Transshipment Revenue Tunning
        unitTransshipmentRevenue = [20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
                                    20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
                                    20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
                                    20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
                                    20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50]
        
        unitTransshipmentCost = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                                1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                                2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2,
                                3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3,
                                4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4]
        
    elif expType == 1:
        # For Holding Cost Sensitivity
        unitTransshipmentRevenue = [38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38,
                                    38, 38, 38, 38, 38, 38, 38]
        unitTransshipmentCost = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2]
    
    
        
    elif expType == 3:
        # For Performance Evaluation
        unitTransshipmentRevenue = [38, 38, 38]
        unitTransshipmentCost = [2, 2, 2]
        
    elif expType == 4:
        # For Debugging
        unitTransshipmentRevenue = [38]
        unitTransshipmentCost = [2]
        
        
    elif expType == 5:
        # For Base stock policy with GA
        # unitTransshipmentRevenue = [38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38, 38]
        # unitTransshipmentCost = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2]
        unitTransshipmentRevenue = [38]
        unitTransshipmentCost = [2]
    
    ############################
    # Unit Holding cost
    ############################
    if expType == 0:
        # For Transshipment Revenue Tuning
        unitHoldingCost = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                            1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                            1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                            1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                            1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
    
    elif expType == 1:
        # For Holding Cost Sensitivity
        unitHoldingCost = [0, 0, 1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6, 7, 7, 8, 8, 9, 9, 10, 10, 11, 11, 12, 12, 13, 13, 14, 14, 15, 15, 16, 16, 17, 17, 18, 18, 19, 19, 20, 20]
    
    
    
    elif expType == 3:
        # For Performance Evaluation
        unitHoldingCost = [1, 1, 1]
    
    elif expType == 4:
        # For Debugging
        unitHoldingCost = [1]
        
    elif expType == 5:
        # For Base stock policy with GA
        # unitHoldingCost = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
        unitHoldingCost = [1]
    
    
    ############################
    # Unit Unmet Demand Penalty
    ############################
    if expType == 0:
        # For Transshipment Revenue Tuning
        unitUnmetDemandPenalty = [26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26,
                                26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26,
                                26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26,
                                26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26,
                                26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26]
    
    elif expType == 1:
        # For Holding Cost Sensitivity
        unitUnmetDemandPenalty = [26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 
                                  26, 26, 26, 26, 26, 26, 26, 26, 26, 26]
    
    
    
    elif expType == 3:
        # For Performance Evaluation
        unitUnmetDemandPenalty = [26, 26, 26]
    
    elif expType == 4:
        # For Debugging
        unitUnmetDemandPenalty = [26]
    
    elif expType == 5:
        # Base stock policy with GA
        # unitUnmetDemandPenalty = [26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26, 26]
        unitUnmetDemandPenalty = [26]
    
    ############################
    # Algorithm type
    ############################
    if expType == 0:
        # For Transshipment Revenue Tuning
        algoType = [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
                    5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
                    5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
                    5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
                    5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]
    
    elif expType == 1:
        # For Holding Cost Sensitivity
        algoType = [5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1, 5, 1]
    
    
    elif expType == 3:
        # For Performance Evaluation
        algoType = [1, 3, 5]
    
    
    elif expType == 4:
        # For Debugging
        algoType = [type]
    
    elif expType == 5:
        # Base stock policy with GA
        # algoType = [type, type, type, type, type, type, type, type, type, type, type, type, type, type, type, type]
        algoType = [type]
    
    
    ############################
    # CR, MR
    ############################    
    if expType == 5:
        # Base stock policy with GA
        # CRList = [0.7, 0.7, 0.7, 0.7, 0.8, 0.8, 0.8, 0.8, 0.9, 0.9, 0.9, 0.9, 1.0, 1.0, 1.0, 1.0]    # 0.7, 0.8, 0.9, 1.0
        # MRList = [0.05, 0.1, 0.15, 0.2, 0.05, 0.1, 0.15, 0.2, 0.05, 0.1, 0.15, 0.2, 0.05, 0.1, 0.15, 0.2]   # 0.05, 0.1, 0.15, 0.2   
        CRList = [1.0]    # 0.7, 0.8, 0.9, 1.0
        MRList = [0.05]   # 0.05, 0.1, 0.15, 0.2   
    
    
    # Loop setting
    maxLoop = len(unitTransshipmentRevenue)
    indexLoop = 0
    
    
    
    # Loop start
    while indexLoop < maxLoop:
        
        # Algorithm setting
        type = algoType[indexLoop]
        
        
        # for training
        rndEGreedy = random.Random(seedEGreedy)
        rndE2Greedy = random.Random(seedE2Greedy)
        rndAction1 = random.Random(seedA1)
        rndAction2 = random.Random(seedA2)
        rndX1 = random.Random(seedX1)
        rndX2 = random.Random(seedX2)
        rndActionO = random.Random(seedO)
        rndPi = random.Random(seedPi)
        
        rndCR1 = random.Random(seedCR1)
        rndCR2 = random.Random(seedCR2)
        rndMR1 = random.Random(seedMR1)
        rndMR2 = random.Random(seedMR2)
        rndOne1 = random.Random(seedOne1)
        rndOne2 = random.Random(seedOne2)
        rndMut1 = random.Random(seedMut1)
        rndMut2 = random.Random(seedMut2)
        rndInitialPop1 = random.Random(seedInitialPop1)
        rndInitialPop2 = random.Random(seedInitialPop2)
        rndMating1 = random.Random(seedMating1)
        rndMating2 = random.Random(seedMating2)
        


        demand1 = Demand(seedDemandMean_1, seedDemandStdDev_1, seedDemandRealization_1, maxT, demandMeanMin_1, demandMeanMax_1, demandStdMin_1, demandStdMax_1)
        demand2 = Demand(seedDemandMean_2, seedDemandStdDev_2, seedDemandRealization_2, maxT, demandMeanMin_2, demandMeanMax_2, demandStdMin_2, demandStdMax_2)


        supplyLoss1 = SupplyLoss(seedSupplyLossMean_1, seedSupplyLossStdDev_1, seedSupplyLossRealization_1, maxT, supplyLossMeanMin_1, supplyLossMeanMax_1, supplyLossStdMin_1, supplyLossStdMax_1)
        supplyLoss2 = SupplyLoss(seedSupplyLossMean_2, seedSupplyLossStdDev_2, seedSupplyLossRealization_2, maxT, supplyLossMeanMin_2, supplyLossMeanMax_2, supplyLossStdMin_2, supplyLossStdMax_2)
        
            
        a1 = Action_retailer()
        a2 = Action_retailer()
        x1 = X_retailer()
        x2 = X_retailer()
        o = Action_parent()
        s = States()
        

        retailer1 = Retailer(initialInventoryLevel_1, len(s.state), len(s.state), len(a1.action), len(o.action), len(x1.action), type, 
                             unitTransshipmentRevenue[indexLoop], unitTransshipmentCost[indexLoop], unitHoldingCost[indexLoop], unitUnmetDemandPenalty[indexLoop])
        retailer2 = Retailer(initialInventoryLevel_2, len(s.state), len(s.state), len(a2.action), len(o.action), len(x2.action), type, 
                             unitTransshipmentRevenue[indexLoop], unitTransshipmentCost[indexLoop], unitHoldingCost[indexLoop], unitUnmetDemandPenalty[indexLoop])
        parent = Parent(len(s.state), len(s.state), len(a1.action), len(a2.action), len(o.action), unitTransshipmentRevenue[indexLoop], unitTransshipmentCost[indexLoop])
        
        
        
        
        
        
        # for test
        rndPi_test = random.Random(seedPi_test)
        demand1_test = Demand(seedDemandMean_1_test, seedDemandStdDev_1_test, seedDemandRealization_1_test, maxT, demandMeanMin_1, demandMeanMax_1, demandStdMin_1, demandStdMax_1)
        demand2_test = Demand(seedDemandMean_2_test, seedDemandStdDev_2_test, seedDemandRealization_2_test, maxT, demandMeanMin_2, demandMeanMax_2, demandStdMin_2, demandStdMax_2)

        supplyLoss1_test = SupplyLoss(seedSupplyLossMean_1_test, seedSupplyLossStdDev_1_test, seedSupplyLossRealization_1_test, maxT, supplyLossMeanMin_1, supplyLossMeanMax_1, supplyLossStdMin_1, supplyLossStdMax_1)
        supplyLoss2_test = SupplyLoss(seedSupplyLossMean_2_test, seedSupplyLossStdDev_2_test, seedSupplyLossRealization_2_test, maxT, supplyLossMeanMin_2, supplyLossMeanMax_2, supplyLossStdMin_2, supplyLossStdMax_2)
        
        
        # for print
        avgRewardList = []
        cpuTimeList = []
        avgUnmetDemandList = []
        avgAmountTransTo1List = []
        avgAmountTransTo2List = []
        avgAmountOrder1List = []
        avgAmountOrder2List = []
        avgOnHandInventory1List = []
        avgOnHandInventory2List = []
        avgUnvisitedStateActionPairCnt = []
        avgUnvisitedStateActionPairCnt_retailer1 = []
        avgUnvisitedStateActionPairCnt_retailer2 = []
        avgUnvisitedStateActionPairCnt_parent = []
        
        rewardList1_LastEpisode = []
        unmetDemandList1_LastEpisode = []
        orderAmount1_LastEpisode = []
        
        rewardList2_LastEpisode = []
        unmetDemandList2_LastEpisode = []
        orderAmount2_LastEpisode = []
        
        # for graph
        avgRewardList = []
        
        
        # Input parameter printing
        now = datetime.now()
        wb = Workbook()
        ws = wb.active
        ws.title = "InputParameter"
        
        if expType == 5:
            xlsFileName = "InputParameter_" + str(CRList[indexLoop]) + "_" + str(MRList[indexLoop]) + ".xlsx"
        else:
            xlsFileName = "InputParameter_" + str(indexLoop) + ".xlsx"
        
        ws["A1"] = "maxT" 
        ws["A2"] = "initialInventoryLevel_1"
        ws["A3"] = "initialInventoryLevel_2"
        ws["A4"] = "demandMeanMin_1"
        ws["A5"] = "demandMeanMax_1" 
        ws["A6"] = "demandMeanMin_2" 
        ws["A7"] = "demandMeanMax_2" 
        ws["A8"] = "supplyLossMeanMin_1" 
        ws["A9"] = "supplyLossMeanMax_1" 
        ws["A10"] = "supplyLossMeanMin_2" 
        ws["A11"] = "supplyLossMeanMax_2" 
        ws["A12"] = "maxEpisode" 
        ws["A13"] = "type" 
        ws["A14"] = "riskSensitivity1" 
        ws["A15"] = "riskSensitivity2" 
        ws["A16"] = "seedDemandMean_1" 
        ws["A17"] = "seedDemandMean_2" 
        ws["A18"] = "seedDemandStdDev_1" 
        ws["A19"] = "seedDemandStdDev_2" 
        ws["A20"] = "seedDemandRealization_1" 
        ws["A21"] = "seedDemandRealization_2" 
        ws["A22"] = "seedSupplyLossMean_1" 
        ws["A23"] = "seedSupplyLossMean_2" 
        ws["A24"] = "seedSupplyLossStdDev_1" 
        ws["A25"] = "seedSupplyLossStdDev_2" 
        ws["A26"] = "seedSupplyLossRealization_1" 
        ws["A27"] = "seedSupplyLossRealization_2" 
        ws["A28"] = "seedEGreedy" 
        ws["A29"] = "seedE2Greedy" 
        ws["A30"] = "seedPi" 
        ws["A31"] = "seedA1" 
        ws["A32"] = "seedA2" 
        ws["A33"] = "seedX1" 
        ws["A34"] = "seedX2" 
        ws["A35"] = "seedO" 
        ws["A36"] = "seedDemandMean_1_test" 
        ws["A37"] = "seedDemandMean_2_test" 
        ws["A38"] = "seedDemandStdDev_1_test" 
        ws["A39"] = "seedDemandStdDev_2_test" 
        ws["A40"] = "seedDemandRealization_1_test" 
        ws["A41"] = "seedDemandRealization_2_test" 
        ws["A42"] = "seedSupplyLossMean_1_test" 
        ws["A43"] = "seedSupplyLossMean_2_test" 
        ws["A44"] = "seedSupplyLossStdDev_1_test" 
        ws["A45"] = "seedSupplyLossStdDev_2_test" 
        ws["A46"] = "seedSupplyLossRealization_1_test" 
        ws["A47"] = "seedSupplyLossRealization_2_test" 
        ws["A48"] = "seedPi_test" 
        
        ws["B1"] = maxT
        ws["B2"] = initialInventoryLevel_1
        ws["B3"] = initialInventoryLevel_2
        ws["B4"] = demandMeanMin_1
        ws["B5"] = demandMeanMax_1 
        ws["B6"] = demandMeanMin_2 
        ws["B7"] = demandMeanMax_2 
        ws["B8"] = supplyLossMeanMin_1
        ws["B9"] = supplyLossMeanMax_1
        ws["B10"] = supplyLossMeanMin_2 
        ws["B11"] = supplyLossMeanMax_2 
        ws["B12"] = maxEpisode
        ws["B13"] = type
        ws["B14"] = riskSensitivity1
        ws["B15"] = riskSensitivity2 
        ws["B16"] = seedDemandMean_1 
        ws["B17"] = seedDemandMean_2 
        ws["B18"] = seedDemandStdDev_1
        ws["B19"] = seedDemandStdDev_2 
        ws["B20"] = seedDemandRealization_1
        ws["B21"] = seedDemandRealization_2 
        ws["B22"] = seedSupplyLossMean_1
        ws["B23"] = seedSupplyLossMean_2 
        ws["B24"] = seedSupplyLossStdDev_1
        ws["B25"] = seedSupplyLossStdDev_2 
        ws["B26"] = seedSupplyLossRealization_1
        ws["B27"] = seedSupplyLossRealization_2 
        ws["B28"] = seedEGreedy
        ws["B29"] = seedE2Greedy 
        ws["B30"] = seedPi
        ws["B31"] = seedA1 
        ws["B32"] = seedA2 
        ws["B33"] = seedX1 
        ws["B34"] = seedX2 
        ws["B35"] = seedO 
        ws["B36"] = seedDemandMean_1_test
        ws["B37"] = seedDemandMean_2_test 
        ws["B38"] = seedDemandStdDev_1_test
        ws["B39"] = seedDemandStdDev_2_test 
        ws["B40"] = seedDemandRealization_1_test
        ws["B41"] = seedDemandRealization_2_test 
        ws["B42"] = seedSupplyLossMean_1_test
        ws["B43"] = seedSupplyLossMean_2_test
        ws["B44"] = seedSupplyLossStdDev_1_test 
        ws["B45"] = seedSupplyLossStdDev_2_test 
        ws["B46"] = seedSupplyLossRealization_1_test
        ws["B47"] = seedSupplyLossRealization_2_test 
        ws["B48"] = seedPi_test
        
        wb.save(xlsFileName)
        
        
        
        
        
        sim = Sim(type, len(s.state), len(s.state), len(a1.action), len(a2.action), len(x1.action), len(x2.action), len(o.action))
        
        
        
        
        # Initialize pi
        
        
        if type == 1:   # Initialize pi(s, a1), pi(s, a2), decentralized q-learning
            sim.InitializePi_decentralized(len(s.state), len(s.state), len(a1.action), len(a2.action))
        
            
        if type == 3:   # Initialize pi(s1, s2, a1, a2, o), centralized q-learning New
            sim.InitializePi_centralized_New(len(s.state), len(s.state), len(a1.action), len(a2.action), len(o.action))
        
        if type == 5 or type == 6 or type == 7:   # Initialize pi(s1, s2, a1, o)
            sim.InitializePi_balancedGap_New(len(s.state), len(s.state), len(a1.action), len(a2.action), len(o.action))
        
        
        
        
        # Learning
        episode = 0
        
        
        timeSecList = []
        for episode in tqdm(range(maxEpisode)):
            
            
                
            if type == 1:   # decentralized q-learning (Trans X)
                start = time.time()
                sim.Run_decentralized(maxT, demand1, demand2, supplyLoss1, supplyLoss2, retailer1, retailer2, parent, a1, a2, s, 
                        round(rndEGreedy.random() * 1000000.0), round(rndAction1.random() * 1000000.0), round(rndAction2.random() * 1000000.0), round(rndPi.random() * 1000000.0), 
                        initialInventoryLevel_1, initialInventoryLevel_2, riskSensitivity1, riskSensitivity2)
                end = time.time()
                timeSecList.append(end - start)
            
            if type == 3:   # centralized q-learning new (Trans O)
                start = time.time()
                sim.Run_centralized_New(maxT, demand1, demand2, supplyLoss1, supplyLoss2, retailer1, retailer2, parent, a1, a2, o, s, 
                        round(rndEGreedy.random() * 1000000.0), round(rndAction1.random() * 1000000.0), round(rndAction2.random() * 1000000.0), 
                        round(rndActionO.random() * 1000000.0), round(rndPi.random() * 1000000.0), initialInventoryLevel_1, initialInventoryLevel_2,
                        riskSensitivity1, riskSensitivity2)    
                end = time.time()
                timeSecList.append(end - start)
            
            if type == 5:   # balanced gap q-learning new (Trans O)
                start = time.time()
                sim.Run_balancedGap_New(maxT, demand1, demand2, supplyLoss1, supplyLoss2, retailer1, retailer2, parent, a1, a2, o, s, 
                        round(rndEGreedy.random() * 1000000.0), round(rndAction1.random() * 1000000.0), round(rndAction2.random() * 1000000.0), 
                        round(rndActionO.random() * 1000000.0), round(rndPi.random() * 1000000.0), initialInventoryLevel_1, initialInventoryLevel_2,
                        riskSensitivity1, riskSensitivity2)
                end = time.time()
                timeSecList.append(end - start)
            
            
            
            # print(f"{end - start:.5f} sec")
            
            
            
            # for test     
            retailer1_test = Retailer(initialInventoryLevel_1, len(s.state), len(s.state), len(a1.action), len(o.action), len(x1.action), type, 
                                      unitTransshipmentRevenue[indexLoop], unitTransshipmentCost[indexLoop], unitHoldingCost[indexLoop], unitUnmetDemandPenalty[indexLoop])
            retailer2_test = Retailer(initialInventoryLevel_2, len(s.state), len(s.state), len(a2.action), len(o.action), len(x2.action), type, 
                                      unitTransshipmentRevenue[indexLoop], unitTransshipmentCost[indexLoop], unitHoldingCost[indexLoop], unitUnmetDemandPenalty[indexLoop])
            
            parent_test = Parent(len(s.state), len(s.state), len(a1.action), len(a2.action), len(o.action), unitTransshipmentRevenue[indexLoop], unitTransshipmentCost[indexLoop])

            
            
            if type == 1:   # decentralized q-learning (Trans X)
                sim.Test_decentralized(maxT, demand1_test, demand2_test, supplyLoss1_test, supplyLoss2_test, retailer1_test, retailer2_test, a1, a2, s, round(rndPi_test.random() * 1000000.0), initialInventoryLevel_1, initialInventoryLevel_2)
            
            
            if type == 3:   # centralized new
                sim.Test_centralized_New(maxT, demand1_test, demand2_test, supplyLoss1_test, supplyLoss2_test, retailer1_test, retailer2_test, parent_test, a1, a2, o, s, round(rndPi_test.random() * 1000000.0), initialInventoryLevel_1, initialInventoryLevel_2)
                
            if type == 5:   # balanced gap q-learning (Trans O)
                sim.Test_balancedGap_New(maxT, demand1_test, demand2_test, supplyLoss1_test, supplyLoss2_test, retailer1_test, retailer2_test, parent_test, a1, a2, o, s, round(rndPi_test.random() * 1000000.0), initialInventoryLevel_1, initialInventoryLevel_2)
            
            if type == 6:   # Base stock with GA
                start = time.time()
                sim.Run_baseStockGA(maxT, demand1_test, demand2_test, supplyLoss1_test, supplyLoss2_test, retailer1_test, retailer2_test, initialInventoryLevel_1, initialInventoryLevel_2, 
                                    baseStockUpperBound, baseStockLowerBound1, baseStockLowerBound2, populationSize, generationSize, supplierCapacity,
                                    round(rndCR1.random() * 1000000.0), round(rndCR2.random() * 1000000.0), round(rndMR1.random() * 1000000.0), round(rndMR2.random() * 1000000.0), 
                                    round(rndOne1.random() * 1000000.0), round(rndOne2.random() * 1000000.0), round(rndMut1.random() * 1000000.0), round(rndMut2.random() * 1000000.0), 
                                    round(rndInitialPop1.random() * 1000000.0), round(rndInitialPop2.random() * 1000000.0), round(rndMating1.random() * 1000000.0), round(rndMating2.random() * 1000000.0), 
                                    CRList[indexLoop], MRList[indexLoop])
                end = time.time()
            timeSecList.append(end - start)
                
            # save for total export out
            totalReward = 0.0
            totalGap = 0.0
            totalUnmetDemand = 0.0
            totalAmountTransTo1 = 0.0
            totalAmountTransTo2 = 0.0
            totalAmountOrder1 = 0.0
            totalAmountOrder2 = 0.0
            totalOnHandInventory1 = 0.0
            totalOnHandInventory2 = 0.0
            
            
            for t in range(maxT):
                # save for total export out
                totalReward += ((retailer1_test.realizedRewardList[t] + retailer2_test.realizedRewardList[t]))
                totalUnmetDemand += ((retailer1_test.unmetDemandList[t] + retailer2_test.unmetDemandList[t]))
                totalAmountTransTo1 += (retailer1_test.amountTransToList[t])
                totalAmountTransTo2 += (retailer2_test.amountTransToList[t])
                totalAmountOrder1 += retailer1_test.orderAmountList[t]
                totalAmountOrder2 += retailer2_test.orderAmountList[t]
                totalOnHandInventory1 += retailer1_test.inventoryLevelBeginningList[t]
                totalOnHandInventory2 += retailer2_test.inventoryLevelBeginningList[t]
                
                # for displaying resilience indicators over periods in the last episode
                if episode == (maxEpisode - 1):
                    rewardList1_LastEpisode.append(retailer1_test.realizedRewardList[t])
                    unmetDemandList1_LastEpisode.append(retailer1_test.unmetDemandList[t])
                    orderAmount1_LastEpisode.append(retailer1_test.orderAmountList[t])
                    
                    rewardList2_LastEpisode.append(retailer2_test.realizedRewardList[t])
                    unmetDemandList2_LastEpisode.append(retailer2_test.unmetDemandList[t])
                    orderAmount2_LastEpisode.append(retailer2_test.orderAmountList[t])
                
                    
            
            # unvisited (state-pair) count sum for each episode
            if expType == 4:
                if type == 1:
                    totalUnvisitCnt1 = len(retailer1.q_cnt[retailer1.q_cnt == 0])
                    avgUnvisitedStateActionPairCnt_retailer1.append(totalUnvisitCnt1)
                        
                    totalUnvisitCnt2 = len(retailer2.q_cnt[retailer2.q_cnt == 0])
                    avgUnvisitedStateActionPairCnt_retailer2.append(totalUnvisitCnt2)
                
                if type == 3:
                    totalUnvisitCnt = len(sim.q_cnt[sim.q_cnt == 0])
                    avgUnvisitedStateActionPairCnt.append(totalUnvisitCnt)
                
                if type == 5:
                    totalUnvisitCnt1 = len(retailer1.q_cnt[retailer1.q_cnt == 0])
                    avgUnvisitedStateActionPairCnt_retailer1.append(totalUnvisitCnt1)
                        
                    totalUnvisitCnt2 = len(retailer2.q_cnt[retailer2.q_cnt == 0])
                    avgUnvisitedStateActionPairCnt_retailer2.append(totalUnvisitCnt2)
                    
                    totalUnvisitCnt3 = len(parent.q_cnt[parent.q_cnt == 0])
                    avgUnvisitedStateActionPairCnt_parent.append(totalUnvisitCnt3)
                
            
            # Save the average (over total period) for each episode
            avgRewardList.append(totalReward / float(maxT))
            avgUnmetDemandList.append(totalUnmetDemand / float(maxT))
            avgAmountTransTo1List.append(totalAmountTransTo1 / float(maxT))
            avgAmountTransTo2List.append(totalAmountTransTo2 / float(maxT))
            avgAmountOrder1List.append(totalAmountOrder1 / float(maxT))
            avgAmountOrder2List.append(totalAmountOrder2 / float(maxT))
            avgOnHandInventory1List.append(totalOnHandInventory1 / float(maxT))
            avgOnHandInventory2List.append(totalOnHandInventory2 / float(maxT))
            
     
        
        if expType == 4:
            
            if type == 1:
                # Display the unvisited (state, action) pair for each episode
                xAxis = range(maxEpisode)
                plt.figure(1)
                plt.plot(xAxis, avgUnvisitedStateActionPairCnt_retailer1, '.', color='b', label='# of unvisited (State, Action) pair of retailer1')
                plt.plot(xAxis, avgUnvisitedStateActionPairCnt_retailer2, '.', color='r', label='# of unvisited (State, Action) pair of retailer2')
                plt.legend()
                plt.xlabel('episode')
                # plt.ylim([300, 1200])
                plt.show()        
                
                
                # unvisited (state, action) pair excel export
                wb = Workbook()
                ws = wb.active
                ws.title = "UnvisitedCnt"
                xlsFileName = "UnvisitedCnt_retailer1.xlsx"
                
                ws["A1"] = "Episode" 
                ws["B1"] = "UnvisitedCnt"
                
                for i in range(len(avgUnvisitedStateActionPairCnt_retailer1)):
                    ws.cell(row=(i + 2), column=1, value=i)
                    ws.cell(row=(i + 2), column=2, value=avgUnvisitedStateActionPairCnt_retailer1[i])
                
                
                wb.save(xlsFileName)
            

                wb = Workbook()
                ws = wb.active
                ws.title = "UnvisitedCnt"
                xlsFileName = "UnvisitedCnt_retailer2.xlsx"
                
                ws["A1"] = "Episode" 
                ws["B1"] = "UnvisitedCnt"
                
                for i in range(len(avgUnvisitedStateActionPairCnt_retailer2)):
                    ws.cell(row=(i + 2), column=1, value=i)
                    ws.cell(row=(i + 2), column=2, value=avgUnvisitedStateActionPairCnt_retailer2[i])
                
                
                wb.save(xlsFileName)
                
                
                            
            if type == 3:
                # Display the unvisited (state, action) pair for each episode
                xAxis = range(maxEpisode)
                plt.figure(1)
                plt.plot(xAxis, avgUnvisitedStateActionPairCnt, '.', color='b', label='# of unvisited (State, Action) pair')
                plt.legend()
                plt.xlabel('episode')
                # plt.ylim([300, 1200])
                plt.show()        
                
                
                # unvisited (state, action) pair excel export
                wb = Workbook()
                ws = wb.active
                ws.title = "UnvisitedCnt"
                xlsFileName = "UnvisitedCnt_" + str(indexLoop) + ".xlsx"
                
                ws["A1"] = "Episode" 
                ws["B1"] = "UnvisitedCnt"
                
                for i in range(len(avgUnvisitedStateActionPairCnt)):
                    ws.cell(row=(i + 2), column=1, value=i)
                    ws.cell(row=(i + 2), column=2, value=avgUnvisitedStateActionPairCnt[i])
                
                
                wb.save(xlsFileName)
            
            
            
            if type == 5:
                # Display the unvisited (state, action) pair for each episode
                xAxis = range(maxEpisode)
                plt.figure(1)
                plt.plot(xAxis, avgUnvisitedStateActionPairCnt_retailer1, '.', color='b', label='# of unvisited (State, Action) pair of retailer1')
                plt.plot(xAxis, avgUnvisitedStateActionPairCnt_retailer2, '.', color='r', label='# of unvisited (State, Action) pair of retailer2')
                plt.plot(xAxis, avgUnvisitedStateActionPairCnt_parent, '.', color='g', label='# of unvisited (State, Action) pair of parent')
                plt.legend()
                plt.xlabel('episode')
                # plt.ylim([300, 1200])
                plt.show()        
                
                
                # unvisited (state, action) pair excel export
                wb = Workbook()
                ws = wb.active
                ws.title = "UnvisitedCnt"
                xlsFileName = "UnvisitedCnt_retailer1.xlsx"
                
                ws["A1"] = "Episode" 
                ws["B1"] = "UnvisitedCnt"
                
                for i in range(len(avgUnvisitedStateActionPairCnt_retailer1)):
                    ws.cell(row=(i + 2), column=1, value=i)
                    ws.cell(row=(i + 2), column=2, value=avgUnvisitedStateActionPairCnt_retailer1[i])
                
                
                wb.save(xlsFileName)
            

                wb = Workbook()
                ws = wb.active
                ws.title = "UnvisitedCnt"
                xlsFileName = "UnvisitedCnt_retailer2.xlsx"
                
                ws["A1"] = "Episode" 
                ws["B1"] = "UnvisitedCnt"
                
                for i in range(len(avgUnvisitedStateActionPairCnt_retailer2)):
                    ws.cell(row=(i + 2), column=1, value=i)
                    ws.cell(row=(i + 2), column=2, value=avgUnvisitedStateActionPairCnt_retailer2[i])
                
                
                wb.save(xlsFileName)
                
                
                
                wb = Workbook()
                ws = wb.active
                ws.title = "UnvisitedCnt"
                xlsFileName = "UnvisitedCnt_parent.xlsx"
                
                ws["A1"] = "Episode" 
                ws["B1"] = "UnvisitedCnt"
                
                for i in range(len(avgUnvisitedStateActionPairCnt_parent)):
                    ws.cell(row=(i + 2), column=1, value=i)
                    ws.cell(row=(i + 2), column=2, value=avgUnvisitedStateActionPairCnt_parent[i])
                
                
                wb.save(xlsFileName)
            
        
        
        # excel export
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        if expType == 5:
            xlsFileName = "Result_" + str(CRList[indexLoop]) + "_" + str(MRList[indexLoop]) + ".xlsx"
        else:
            xlsFileName = "Result_" + str(indexLoop) + ".xlsx"
        
        ws["A1"] = "Algorithm" 
        ws["B1"] = "Episode" 
        ws["C1"] = "AvgProfit(1&2sum)"
        ws["D1"] = "AvgUnmetDemand"
        ws["E1"] = "AvgAmountTransTo1"
        ws["F1"] = "AvgAmountTransTo2"
        ws["G1"] = "SumAmountTransTo1"
        ws["H1"] = "SumAmountTransTo2"
        ws["I1"] = "AvgAmountOrder1"
        ws["J1"] = "AvgAmountOrder2"
        ws["K1"] = "AvgOnHandInventory1"
        ws["L1"] = "AvgOnHandInventory2"
        ws["M1"] = "Time(sec)"
        ws["N1"] = "AvgAmountTransTo(1&2sum)"
        ws["O1"] = "AvgAmountOrder(1&2sum)"
        ws["P1"] = "AvgOnHandInventory(1&2sum)"
        
        
        for i in range(len(avgRewardList)):
            
            if expType == 5:
                ws.cell(row=(i + 2), column=1, value="GA")
            else:
                if type == 5:
                    ws.cell(row=(i + 2), column=1, value="Hetero-Maximax")
                elif type == 1:
                    ws.cell(row=(i + 2), column=1, value="Decentralized")
                elif type == 3:
                    ws.cell(row=(i + 2), column=1, value="Centralized")
            
            
            ws.cell(row=(i + 2), column=2, value=i)
            ws.cell(row=(i + 2), column=3, value=avgRewardList[i])
            ws.cell(row=(i + 2), column=4, value=avgUnmetDemandList[i])
            ws.cell(row=(i + 2), column=5, value=avgAmountTransTo1List[i])
            ws.cell(row=(i + 2), column=6, value=avgAmountTransTo2List[i])
            ws.cell(row=(i + 2), column=7, value=(avgAmountTransTo1List[i] * maxT))
            ws.cell(row=(i + 2), column=8, value=(avgAmountTransTo2List[i] * maxT))
            ws.cell(row=(i + 2), column=9, value=avgAmountOrder1List[i])
            ws.cell(row=(i + 2), column=10, value=avgAmountOrder2List[i])
            ws.cell(row=(i + 2), column=11, value=avgOnHandInventory1List[i])
            ws.cell(row=(i + 2), column=12, value=avgOnHandInventory2List[i])
            ws.cell(row=(i + 2), column=13, value=timeSecList[i]) 
            ws.cell(row=(i + 2), column=14, value=((avgAmountTransTo1List[i] + avgAmountTransTo2List[i]))) 
            ws.cell(row=(i + 2), column=15, value=((avgAmountOrder1List[i] + avgAmountOrder2List[i]))) 
            ws.cell(row=(i + 2), column=16, value=((avgOnHandInventory1List[i] + avgOnHandInventory2List[i]))) 
        
        
        wb.save(xlsFileName)
        
        
        
        # excel export for SC resilience indicators
        wb = Workbook()
        ws = wb.active
        ws.title = "Result"
        if expType == 5:
            xlsFileName = "Result_SCResilience_" + str(CRList[indexLoop]) + "_" + str(MRList[indexLoop]) + ".xlsx"
        else:
            xlsFileName = "Result_SCResilience_" + str(indexLoop) + ".xlsx"
        
        ws["A1"] = "Algorithm" 
        ws["B1"] = "Episode" 
        ws["C1"] = "Period" 
        ws["D1"] = "Demand1"
        ws["E1"] = "Demand2"
        ws["F1"] = "FillRate1"
        ws["G1"] = "FillRate2"
        ws["H1"] = "Loss1"
        ws["I1"] = "Loss2"
        ws["J1"] = "Loss(1&2sum)"
        ws["K1"] = "UnmetDemand1"
        ws["L1"] = "UnmetDemand2"
        ws["M1"] = "UnmetDemand(1&2sum)"
        
        
        for i in range(maxT):    
            if expType == 5:
                ws.cell(row=(i + 2), column=1, value="GA")
            else:
                if type == 5:
                    ws.cell(row=(i + 2), column=1, value="Hetero-Maximax")
                elif type == 1:
                    ws.cell(row=(i + 2), column=1, value="Decentralized")
                elif type == 3:
                    ws.cell(row=(i + 2), column=1, value="Centralized")
            
            
            ws.cell(row=(i + 2), column=2, value=maxEpisode)
            ws.cell(row=(i + 2), column=3, value=(i+1))
            ws.cell(row=(i + 2), column=4, value=demand1_test.realizedArray[i])
            ws.cell(row=(i + 2), column=5, value=demand2_test.realizedArray[i])
            if retailer1_test.orderAmountList[i] == 0:
                ws.cell(row=(i + 2), column=6, value=1)
            else:
                ws.cell(row=(i + 2), column=6, value=(max(retailer1_test.orderAmountList[i] - supplyLoss1_test.realizedArray[i], 0) / retailer1_test.orderAmountList[i]))
            
            if retailer2_test.orderAmountList[i] == 0:
                ws.cell(row=(i + 2), column=7, value=1)
            else:
                ws.cell(row=(i + 2), column=7, value=(max(retailer2_test.orderAmountList[i] - supplyLoss2_test.realizedArray[i], 0) / retailer2_test.orderAmountList[i]))
            
            
            loss1 = 0
            if retailer1_test.realizedRewardList[i] < 0:
                loss1 = -retailer1_test.realizedRewardList[i]
            else:
                loss1 = 0
            
            loss2 = 0
            if retailer2_test.realizedRewardList[i] < 0:
                loss2 = -retailer2_test.realizedRewardList[i]
            else:
                loss2 = 0
                

            ws.cell(row=(i + 2), column=8, value=loss1)
            ws.cell(row=(i + 2), column=9, value=loss2)

            ws.cell(row=(i + 2), column=10, value=(loss1+loss2))

            ws.cell(row=(i + 2), column=11, value=retailer1_test.unmetDemandList[i])
            ws.cell(row=(i + 2), column=12, value=retailer2_test.unmetDemandList[i])
            ws.cell(row=(i + 2), column=13, value=(retailer1_test.unmetDemandList[i]+retailer2_test.unmetDemandList[i]))
        
        wb.save(xlsFileName)
        
        
         
        
        # Calculate the average for results after episode 50
        avgReward_sum = 0.0
        avgUnmetDemand_sum = 0.0
        avgAmountTransTo1_sum = 0.0
        avgAmountTransTo2_sum = 0.0
        avgAmountOrder1_sum = 0.0
        avgAmountOrder2_sum = 0.0
        avgOnHandInventory1_sum = 0.0
        avgOnHandInventory2_sum = 0.0
        avgTimeSec_sum = 0.0
        
        for i in range(50, len(avgRewardList)):
            avgReward_sum = (avgReward_sum + avgRewardList[i])
            avgUnmetDemand_sum = (avgUnmetDemand_sum + avgUnmetDemandList[i])
            avgAmountTransTo1_sum = (avgAmountTransTo1_sum + avgAmountTransTo1List[i])
            avgAmountTransTo2_sum = (avgAmountTransTo2_sum + avgAmountTransTo2List[i])
            avgAmountOrder1_sum = (avgAmountOrder1_sum + avgAmountOrder1List[i])
            avgAmountOrder2_sum = (avgAmountOrder2_sum + avgAmountOrder2List[i])
            avgOnHandInventory1_sum = (avgOnHandInventory1_sum + avgOnHandInventory1List[i])
            avgOnHandInventory2_sum = (avgOnHandInventory2_sum + avgOnHandInventory2List[i])
            avgTimeSec_sum = (avgTimeSec_sum + timeSecList[i])
            
        avgAvgRewardList.append(avgReward_sum / float(len(avgRewardList) - 50)) 
        avgAvgUnmetDemandList.append(avgUnmetDemand_sum / float(len(avgRewardList) - 50)) 
        avgAvgAmountTransTo1List.append(avgAmountTransTo1_sum / float(len(avgRewardList) - 50)) 
        avgAvgAmountTransTo2List.append(avgAmountTransTo2_sum / float(len(avgRewardList) - 50)) 
        avgAvgAmountOrder1List.append(avgAmountOrder1_sum / float(len(avgRewardList) - 50)) 
        avgAvgAmountOrder2List.append(avgAmountOrder2_sum / float(len(avgRewardList) - 50)) 
        avgAvgOnHandInventory1List.append(avgOnHandInventory1_sum / float(len(avgRewardList) - 50)) 
        avgAvgOnHandInventory2List.append(avgOnHandInventory2_sum / float(len(avgRewardList) - 50)) 
        avgTimeSecList.append(avgTimeSec_sum / float(len(avgRewardList) - 50))
            
        indexLoop = indexLoop + 1
        
    
    if expType == 0:    
        # TransshipmentRevenue Tuning excel export
        wb = Workbook()
        ws = wb.active
        ws.title = "UnitTransRevenueSensitivity"
        xlsFileName = "UnitTransRevenueSensitivity.xlsx"
        
        ws["A1"] = "No"
        ws["B1"] = "Algorithm" 
        ws["C1"] = "UnitTransRevenue"
        ws["D1"] = "UnitTransCost"
        ws["E1"] = "AvgofAvgAmountTransTo1"
        ws["F1"] = "AvgofAvgAmountTransTo2"
        ws["G1"] = "AvgofAvgOrderAmount1"
        ws["H1"] = "AvgofAvgOrderAmount2"
        ws["I1"] = "AvgofUnmetDemand(1&2sum)"
        ws["J1"] = "AvgofAvgOnHandInventory1"
        ws["K1"] = "AvgofAvgOnHandInventory2"
        ws["L1"] = "AvgofProfit(1&2sum)"
        ws["M1"] = "AvgTotalTime(sec)"
        ws["N1"] = "AvgofAvgAmountTransTo(1&2sum)"
        ws["O1"] = "AvgofAvgOrderAmount(1&2sum)"
        ws["P1"] = "AvgofAvgOnHandInventory(1&2sum)"
        
        for i in range(len(avgAvgRewardList)):
            ws.cell(row=(i + 2), column=1, value=i)
            ws.cell(row=(i + 2), column=2, value="Hetero-Maximax")
            ws.cell(row=(i + 2), column=3, value=unitTransshipmentRevenue[i])
            ws.cell(row=(i + 2), column=4, value=unitTransshipmentCost[i])
            ws.cell(row=(i + 2), column=5, value=avgAvgAmountTransTo1List[i])
            ws.cell(row=(i + 2), column=6, value=avgAvgAmountTransTo2List[i])
            ws.cell(row=(i + 2), column=7, value=avgAvgAmountOrder1List[i])
            ws.cell(row=(i + 2), column=8, value=avgAvgAmountOrder2List[i])
            ws.cell(row=(i + 2), column=9, value=avgAvgUnmetDemandList[i])
            ws.cell(row=(i + 2), column=10, value=avgAvgOnHandInventory1List[i])
            ws.cell(row=(i + 2), column=11, value=avgAvgOnHandInventory2List[i])
            ws.cell(row=(i + 2), column=12, value=avgAvgRewardList[i])
            ws.cell(row=(i + 2), column=13, value=avgTimeSecList[i])
            ws.cell(row=(i + 2), column=14, value=((avgAvgAmountTransTo1List[i] + avgAvgAmountTransTo2List[i]))) 
            ws.cell(row=(i + 2), column=15, value=((avgAvgAmountOrder1List[i] + avgAvgAmountOrder2List[i]))) 
            ws.cell(row=(i + 2), column=16, value=((avgAvgOnHandInventory1List[i] + avgAvgOnHandInventory2List[i]))) 
        
        
        wb.save(xlsFileName)
    
    
    
    elif expType == 1:  
        # HoldingCost Sensitivity excel export
        wb = Workbook()
        ws = wb.active
        ws.title = "UnitHoldingCostSensitivity"
        xlsFileName = "UnitHoldingCostSensitivity.xlsx"
        
        ws["A1"] = "No"
        ws["B1"] = "Algorithm" 
        ws["C1"] = "UnitHoldingCost"
        ws["D1"] = "AvgofAvgAmountTransTo1"
        ws["E1"] = "AvgofAvgAmountTransTo2"
        ws["F1"] = "AvgofAvgOrderAmount1"
        ws["G1"] = "AvgofAvgOrderAmount2"
        ws["H1"] = "AvgofUnmetDemand(1&2sum)"
        ws["I1"] = "AvgofAvgOnHandInventory1"
        ws["J1"] = "AvgofAvgOnHandInventory2"
        ws["K1"] = "AvgofProfit(1&2sum)"
        ws["L1"] = "AvgTotalTime(sec)"
        ws["M1"] = "AvgofAvgAmountTransTo(1&2sum)"
        ws["N1"] = "AvgofAvgOrderAmount(1&2sum)"
        ws["O1"] = "AvgofAvgOnHandInventory(1&2sum)"
        
        for i in range(len(avgAvgRewardList)):
            ws.cell(row=(i + 2), column=1, value=i)
            if algoType[i] == 5:
                ws.cell(row=(i + 2), column=2, value="Hetero-Maximax")
            elif algoType[i] == 1:
                ws.cell(row=(i + 2), column=2, value="Decentralized")
            ws.cell(row=(i + 2), column=3, value=unitHoldingCost[i])
            ws.cell(row=(i + 2), column=4, value=avgAvgAmountTransTo1List[i])
            ws.cell(row=(i + 2), column=5, value=avgAvgAmountTransTo2List[i])
            ws.cell(row=(i + 2), column=6, value=avgAvgAmountOrder1List[i])
            ws.cell(row=(i + 2), column=7, value=avgAvgAmountOrder2List[i])
            ws.cell(row=(i + 2), column=8, value=avgAvgUnmetDemandList[i])
            ws.cell(row=(i + 2), column=9, value=avgAvgOnHandInventory1List[i])
            ws.cell(row=(i + 2), column=10, value=avgAvgOnHandInventory2List[i])
            ws.cell(row=(i + 2), column=11, value=avgAvgRewardList[i])
            ws.cell(row=(i + 2), column=12, value=avgTimeSecList[i])
            ws.cell(row=(i + 2), column=13, value=((avgAvgAmountTransTo1List[i] + avgAvgAmountTransTo2List[i]))) 
            ws.cell(row=(i + 2), column=14, value=((avgAvgAmountOrder1List[i] + avgAvgAmountOrder2List[i]))) 
            ws.cell(row=(i + 2), column=15, value=((avgAvgOnHandInventory1List[i] + avgAvgOnHandInventory2List[i]))) 
            
        
        wb.save(xlsFileName)
    
    
    elif expType == 2:  
        # Unmet Demand Penalty Sensitivity excel export
        wb = Workbook()
        ws = wb.active
        ws.title = "UnitUnmetPenaltySensitivity"
        xlsFileName = "UnitUnmetPenaltySensitivity.xlsx"
        
        ws["A1"] = "No"
        ws["B1"] = "Algorithm" 
        ws["C1"] = "UnitUnmetDemandPenalty"
        ws["D1"] = "AvgofAvgAmountTransTo1"
        ws["E1"] = "AvgofAvgAmountTransTo2"
        ws["F1"] = "AvgofAvgOrderAmount1"
        ws["G1"] = "AvgofAvgOrderAmount2"
        ws["H1"] = "AvgofUnmetDemand(1&2sum)"
        ws["I1"] = "AvgofAvgOnHandInventory1"
        ws["J1"] = "AvgofAvgOnHandInventory2"
        ws["K1"] = "AvgofProfit(1&2sum)"
        ws["L1"] = "AvgTotalTime(sec)"
        ws["M1"] = "AvgofAvgAmountTransTo(1&2sum)"
        ws["N1"] = "AvgofAvgOrderAmount(1&2sum)"
        ws["O1"] = "AvgofAvgOnHandInventory(1&2sum)"
        
        for i in range(len(avgAvgRewardList)):
            ws.cell(row=(i + 2), column=1, value=i)
            if algoType[i] == 5:
                ws.cell(row=(i + 2), column=2, value="Hetero-Maximax")
            elif algoType[i] == 1:
                ws.cell(row=(i + 2), column=2, value="Decentralized")
            ws.cell(row=(i + 2), column=3, value=unitUnmetDemandPenalty[i])
            ws.cell(row=(i + 2), column=4, value=avgAvgAmountTransTo1List[i])
            ws.cell(row=(i + 2), column=5, value=avgAvgAmountTransTo2List[i])
            ws.cell(row=(i + 2), column=6, value=avgAvgAmountOrder1List[i])
            ws.cell(row=(i + 2), column=7, value=avgAvgAmountOrder2List[i])
            ws.cell(row=(i + 2), column=8, value=avgAvgUnmetDemandList[i])
            ws.cell(row=(i + 2), column=9, value=avgAvgOnHandInventory1List[i])
            ws.cell(row=(i + 2), column=10, value=avgAvgOnHandInventory2List[i])
            ws.cell(row=(i + 2), column=11, value=avgAvgRewardList[i])
            ws.cell(row=(i + 2), column=12, value=avgTimeSecList[i])
            ws.cell(row=(i + 2), column=13, value=((avgAvgAmountTransTo1List[i] + avgAvgAmountTransTo2List[i]))) 
            ws.cell(row=(i + 2), column=14, value=((avgAvgAmountOrder1List[i] + avgAvgAmountOrder2List[i]))) 
            ws.cell(row=(i + 2), column=15, value=((avgAvgOnHandInventory1List[i] + avgAvgOnHandInventory2List[i]))) 
        
        
        wb.save(xlsFileName)
    